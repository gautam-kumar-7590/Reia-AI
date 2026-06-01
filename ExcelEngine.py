# ExcelEngine.py — Complete Rewrite (Data-Driven, All Rows, All Charts)
import os
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ─── COLORS ───────────────────────────────────────────────────────────────────
NAVY     = "1a1f2e"
AMBER    = "F0A500"
BLUE     = "00D4FF"
WHITE    = "FFFFFF"
RED      = "FF4444"
GREEN    = "00C851"
GRAY     = "2d3347"
DARKGRAY = "1e2335"
ORANGE   = "FF6B35"

# ─── STYLE HELPERS ────────────────────────────────────────────────────────────
def hdr(ws, row, col, value, bg=NAVY, fg=AMBER, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(color=fg, bold=bold, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def dat(ws, row, col, value, bg=DARKGRAY, fg=WHITE, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(color=fg, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return c

def flag_cell(ws, row, col, value, good=True):
    color = GREEN if good else RED
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=DARKGRAY)
    c.font = Font(color=color, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

def set_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def write_df(ws, df, start_row=2, header_row=1):
    """Write full DataFrame — every row, every column."""
    for ci, col_name in enumerate(df.columns, start=1):
        hdr(ws, header_row, ci, str(col_name))
        ws.column_dimensions[get_column_letter(ci)].width = max(14, min(38, len(str(col_name)) + 6))
    for ri, (_, row_data) in enumerate(df.iterrows(), start=start_row):
        bg = DARKGRAY if ri % 2 == 0 else GRAY
        for ci, val in enumerate(row_data, start=1):
            if isinstance(val, float):
                val = round(val, 2)
            dat(ws, ri, ci, val, bg=bg)
    return start_row + len(df)

# Global flag — set to False for large datasets to save memory
_CHARTS_ENABLED = True

def bar_chart(ws, title, data_ref, cats_ref, anchor, width=20, height=13):
    if not _CHARTS_ENABLED: return
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.grouping = "clustered"
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)

def line_chart(ws, title, data_ref, cats_ref, anchor, width=22, height=13):
    if not _CHARTS_ENABLED: return
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)

def pie_chart(ws, title, data_ref, cats_ref, anchor, width=14, height=12):
    if not _CHARTS_ENABLED: return
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)


# ─── AUTO-DETECT COLUMNS ──────────────────────────────────────────────────────
def detect_columns(df: pd.DataFrame) -> dict:
    # Normalize: lowercase + spaces→underscores for matching
    col_map = {c.lower().strip().replace(" ", "_"): c for c in df.columns}
    result = {}

    patterns = {
        'sales':     ['sales', 'revenue', 'amount', 'total_sales', 'sale_amount', 'turnover', 'value', 'total'],
        'profit':    ['profit', 'net_income', 'earnings', 'net_profit', 'margin_amount', 'gain'],
        'quantity':  ['quantity', 'qty', 'units', 'count', 'volume', 'orders', 'num_orders'],
        'region':    ['region', 'area', 'zone', 'territory', 'market'],
        'category':  ['category', 'sub_category', 'sub-category', 'department', 'division', 'type'],
        'product':   ['product_name', 'product', 'item_name', 'item', 'sku', 'description', 'product_id'],
        'date':      ['order_date', 'date', 'transaction_date', 'period', 'ship_date', 'created_at'],
        'ship_cost': ['shipping_cost', 'freight', 'delivery_cost', 'ship_cost'],
        'ship_mode': ['ship_mode', 'shipping_mode', 'delivery_mode', 'shipment_type'],
        'customer':  ['customer_name', 'customer', 'client', 'buyer', 'name'],
        'segment':   ['customer_segment', 'segment', 'market_segment', 'customer_type'],
        'country':   ['country', 'nation', 'location', 'country_region'],
        'city':      ['city', 'town'],
        'order_id':  ['order_id', 'order_number', 'transaction_id'],
        'discount':  ['discount', 'discount_rate', 'disc'],
    }

    for semantic, keywords in patterns.items():
        for kw in keywords:
            if kw in col_map:
                result[semantic] = col_map[kw]
                break
        if semantic not in result:
            for col_lower, col_orig in col_map.items():
                for kw in keywords:
                    kw_norm = kw.replace(" ", "_")
                    if kw_norm in col_lower or kw_norm.replace("_", "") in col_lower.replace("_", ""):
                        result[semantic] = col_orig
                        break
                if semantic in result:
                    break

    result['_numeric'] = df.select_dtypes(include=[np.number]).columns.tolist()
    result['_categorical'] = df.select_dtypes(include=['object']).columns.tolist()
    return result


# ─── SHEET 1: OVERVIEW ────────────────────────────────────────────────────────
def build_overview(ws, df, company_name, doc_type, cols):
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 30, "B": 22, "C": 22, "D": 22, "E": 22})

    hdr(ws, 1, 1, f"REIA — FINANCIAL REPORT: {company_name.upper()}", size=14, bg=NAVY, fg=AMBER)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 32

    # Meta
    row = 3
    for label, val in [
        ("Report Date",        datetime.now().strftime("%d %B %Y")),
        ("Document Type",      doc_type.upper()),
        ("Total Rows Analyzed",f"{len(df):,}"),
        ("Total Columns",      str(len(df.columns))),
        ("Columns Detected",   ", ".join(list(cols.keys())[:8])),
    ]:
        dat(ws, row, 1, label, bold=True)
        dat(ws, row, 2, val)
        row += 1

    row += 1
    hdr(ws, row, 1, "KEY METRICS", size=12, bg=GRAY, fg=AMBER)
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    kpis = []
    if 'sales' in cols:
        s = df[cols['sales']].sum()
        kpis.append(("Total Sales",   f"${s:,.2f}",  True))
        kpis.append(("Avg Sale",      f"${df[cols['sales']].mean():,.2f}", True))
    if 'profit' in cols:
        p = df[cols['profit']].sum()
        kpis.append(("Total Profit",  f"${p:,.2f}", p > 0))
        losses = (df[cols['profit']] < 0).sum()
        kpis.append(("Loss-Making Rows", str(losses), losses == 0))
        if 'sales' in cols and df[cols['sales']].sum() > 0:
            margin = p / df[cols['sales']].sum() * 100
            kpis.append(("Profit Margin", f"{margin:.1f}%", margin > 10))
    if 'quantity' in cols:
        kpis.append(("Units Sold",    f"{int(df[cols['quantity']].sum()):,}", True))
    if 'ship_cost' in cols:
        kpis.append(("Avg Ship Cost", f"${df[cols['ship_cost']].mean():.2f}", df[cols['ship_cost']].mean() < 40))
    if 'region' in cols:
        kpis.append(("Regions",       str(df[cols['region']].nunique()), True))
    if 'product' in cols:
        kpis.append(("Unique Products", str(df[cols['product']].nunique()), True))
    if 'customer' in cols:
        kpis.append(("Unique Customers", str(df[cols['customer']].nunique()), True))

    per_row = 4
    for i, (label, value, good) in enumerate(kpis):
        c = (i % per_row) + 1
        r = row + (i // per_row) * 3
        hdr(ws, r, c, label, bg=GRAY, fg=WHITE, size=10)
        flag_cell(ws, r + 1, c, value, good=good)

    row += (len(kpis) // per_row + 1) * 3 + 2
    hdr(ws, row, 1, "Generated by Reia CFO — AI Financial Analyst", bg=NAVY, fg=BLUE, size=10)
    ws.merge_cells(f"A{row}:E{row}")


# ─── SHEET 2: RAW DATA (ALL ROWS, ALL COLUMNS) ────────────────────────────────
def build_raw_data(ws, df):
    ws.sheet_view.showGridLines = False
    num_cols = len(df.columns)
    last_col = get_column_letter(num_cols)
    # Cap at 5000 rows to avoid OOM on Streamlit Cloud
    MAX_ROWS = 5000
    df_display = df.head(MAX_ROWS)
    note = f"COMPLETE DATASET — {len(df):,} ROWS × {num_cols} COLUMNS"
    if len(df) > MAX_ROWS:
        note += f" (showing first {MAX_ROWS:,} rows)"
    hdr(ws, 1, 1, note, size=12, bg=NAVY, fg=AMBER)
    ws.merge_cells(f"A1:{last_col}1")
    write_df(ws, df_display, start_row=3, header_row=2)


# ─── SHEET 3: SALES ANALYSIS ──────────────────────────────────────────────────
def build_sales(ws, df, cols):
    ws.sheet_view.showGridLines = False
    hdr(ws, 1, 1, "SALES ANALYSIS", size=13, bg=NAVY, fg=AMBER)
    ws.merge_cells("A1:F1")
    set_widths(ws, {"A": 28, "B": 18, "C": 18, "D": 18})

    if 'sales' not in cols:
        dat(ws, 3, 1, "No sales column detected in dataset")
        return

    s_col = cols['sales']
    row = 3

    def write_group(title, groupby_col, anchor, chart_type="bar"):
        nonlocal row
        hdr(ws, row, 1, title, bg=GRAY)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        grp = df.groupby(groupby_col).agg(
            **{
                'Total Sales': (s_col, 'sum'),
                **({"Total Profit": (cols['profit'], 'sum')} if 'profit' in cols else {}),
                **({"Qty": (cols['quantity'], 'sum')} if 'quantity' in cols else {}),
            }
        ).sort_values('Total Sales', ascending=False).reset_index()
        grp['Total Sales'] = grp['Total Sales'].round(2)

        headers = [groupby_col, 'Total Sales']
        if 'profit' in cols: headers.append('Total Profit')
        if 'quantity' in cols: headers.append('Qty')
        for ci, h in enumerate(headers, 1):
            hdr(ws, row, ci, h)
        row += 1

        data_start = row
        for _, r in grp.iterrows():
            dat(ws, row, 1, r[groupby_col])
            dat(ws, row, 2, r['Total Sales'], align="right")
            col = 3
            if 'profit' in cols and 'Total Profit' in grp.columns:
                flag_cell(ws, row, col, round(r['Total Profit'], 2), good=r['Total Profit'] > 0)
                col += 1
            if 'quantity' in cols and 'Qty' in grp.columns:
                dat(ws, row, col, int(r['Qty']), align="right")
            row += 1
        data_end = row - 1

        d_ref = Reference(ws, min_col=2, min_row=data_start - 1, max_row=data_end)
        c_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        if chart_type == "pie":
            pie_chart(ws, title, d_ref, c_ref, anchor)
        else:
            bar_chart(ws, title, d_ref, c_ref, anchor)
        row += 2

    if 'region' in cols:
        write_group("Sales by Region", cols['region'], f"F{row}")
    if 'category' in cols:
        write_group("Sales by Category", cols['category'], f"F{row}", chart_type="pie")
    if 'segment' in cols:
        write_group("Sales by Customer Segment", cols['segment'], f"F{row}")
    if 'ship_mode' in cols:
        write_group("Sales by Ship Mode", cols['ship_mode'], f"F{row}", chart_type="pie")


# ─── SHEET 4: PRODUCT ANALYSIS ────────────────────────────────────────────────
def build_products(ws, df, cols):
    ws.sheet_view.showGridLines = False
    hdr(ws, 1, 1, "PRODUCT ANALYSIS — TOP & BOTTOM PERFORMERS", size=13, bg=NAVY, fg=AMBER)
    ws.merge_cells("A1:G1")
    set_widths(ws, {"A": 40, "B": 18, "C": 18, "D": 18, "E": 18})

    if 'product' not in cols or 'sales' not in cols:
        dat(ws, 3, 1, "No product/sales column detected")
        return

    p_col = cols['product']
    s_col = cols['sales']

    agg = {'Total Sales': (s_col, 'sum')}
    if 'profit' in cols: agg['Total Profit'] = (cols['profit'], 'sum')
    if 'quantity' in cols: agg['Qty Sold'] = (cols['quantity'], 'sum')

    product_df = df.groupby(p_col).agg(**agg).reset_index()
    product_df['Total Sales'] = product_df['Total Sales'].round(2)

    row = 3

    # Top 20
    top20 = product_df.sort_values('Total Sales', ascending=False).head(20)
    hdr(ws, row, 1, f"TOP 20 PRODUCTS BY SALES (out of {len(product_df)} products)", bg=GREEN, fg=WHITE)
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    headers = [p_col, 'Total Sales']
    if 'profit' in cols: headers.append('Total Profit')
    if 'quantity' in cols: headers.append('Qty Sold')
    for ci, h in enumerate(headers, 1):
        hdr(ws, row, ci, h)
    row += 1

    top_start = row
    for _, r in top20.iterrows():
        dat(ws, row, 1, r[p_col])
        dat(ws, row, 2, r['Total Sales'], align="right")
        ci = 3
        if 'profit' in cols:
            flag_cell(ws, row, ci, round(r['Total Profit'], 2), good=r['Total Profit'] > 0)
            ci += 1
        if 'quantity' in cols:
            dat(ws, row, ci, int(r['Qty Sold']), align="right")
        row += 1
    top_end = row - 1

    d_ref = Reference(ws, min_col=2, min_row=top_start - 1, max_row=top_end)
    c_ref = Reference(ws, min_col=1, min_row=top_start, max_row=top_end)
    bar_chart(ws, "Top 20 Products by Sales", d_ref, c_ref, f"G{top_start - 1}", width=24, height=16)

    row += 2

    # Bottom 10 by profit
    if 'profit' in cols:
        bottom10 = product_df.sort_values('Total Profit').head(10)
        hdr(ws, row, 1, "WORST 10 PRODUCTS BY PROFIT (Biggest Loss Makers)", bg=RED, fg=WHITE)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        for ci, h in enumerate(headers, 1):
            hdr(ws, row, ci, h, bg=GRAY)
        row += 1

        for _, r in bottom10.iterrows():
            dat(ws, row, 1, r[p_col])
            dat(ws, row, 2, r['Total Sales'], align="right")
            ci = 3
            if 'profit' in cols:
                flag_cell(ws, row, ci, round(r['Total Profit'], 2), good=r['Total Profit'] > 0)
                ci += 1
            if 'quantity' in cols:
                dat(ws, row, ci, int(r['Qty Sold']), align="right")
            row += 1


# ─── SHEET 5: PROFIT ANALYSIS ─────────────────────────────────────────────────
def build_profit(ws, df, cols):
    ws.sheet_view.showGridLines = False
    hdr(ws, 1, 1, "PROFIT ANALYSIS", size=13, bg=NAVY, fg=AMBER)
    ws.merge_cells("A1:F1")
    set_widths(ws, {"A": 30, "B": 22, "C": 22, "D": 22})

    if 'profit' not in cols:
        dat(ws, 3, 1, "No profit column detected in dataset")
        return

    p_col = cols['profit']
    row = 3

    total_profit = df[p_col].sum()
    profitable   = (df[p_col] > 0).sum()
    loss_rows    = (df[p_col] < 0).sum()
    total_loss   = df[df[p_col] < 0][p_col].sum()

    hdr(ws, row, 1, "PROFIT SUMMARY", bg=GRAY)
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    summary = [
        ("Total Profit",       f"${total_profit:,.2f}", total_profit > 0),
        ("Profitable Rows",    f"{profitable:,}",       True),
        ("Loss-Making Rows",   f"{loss_rows:,}",        loss_rows == 0),
        ("Total Loss Amount",  f"${total_loss:,.2f}",   total_loss == 0),
    ]
    if 'sales' in cols and df[cols['sales']].sum() > 0:
        margin = total_profit / df[cols['sales']].sum() * 100
        summary.append(("Overall Margin %", f"{margin:.2f}%", margin > 10))

    for label, value, good in summary:
        dat(ws, row, 1, label, bold=True)
        flag_cell(ws, row, 2, value, good=good)
        row += 1

    row += 1

    def profit_by_group(groupby_col, title, anchor):
        nonlocal row
        hdr(ws, row, 1, title, bg=GRAY)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1
        grp = df.groupby(groupby_col)[p_col].sum().sort_values(ascending=False).reset_index()
        hdr(ws, row, 1, groupby_col)
        hdr(ws, row, 2, "Total Profit")
        hdr(ws, row, 3, "Status")
        row += 1
        data_start = row
        for _, r in grp.iterrows():
            dat(ws, row, 1, r[groupby_col])
            flag_cell(ws, row, 2, round(r[p_col], 2), good=r[p_col] > 0)
            flag_cell(ws, row, 3, "✅ Profit" if r[p_col] > 0 else "❌ Loss", good=r[p_col] > 0)
            row += 1
        data_end = row - 1
        d = Reference(ws, min_col=2, min_row=data_start - 1, max_row=data_end)
        c = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        bar_chart(ws, title, d, c, anchor)
        row += 2

    if 'region' in cols:
        profit_by_group(cols['region'], "Profit by Region", f"F{row}")
    if 'category' in cols:
        profit_by_group(cols['category'], "Profit by Category", f"F{row}")
    if 'segment' in cols:
        profit_by_group(cols['segment'], "Profit by Segment", f"F{row}")


# ─── SHEET 6: SHIPPING ANALYSIS ───────────────────────────────────────────────
def build_shipping(ws, df, cols):
    ws.sheet_view.showGridLines = False
    hdr(ws, 1, 1, "SHIPPING ANALYSIS", size=13, bg=NAVY, fg=AMBER)
    ws.merge_cells("A1:F1")
    set_widths(ws, {"A": 28, "B": 20, "C": 20, "D": 20})

    row = 3

    if 'ship_mode' in cols:
        s_col = cols['ship_mode']
        hdr(ws, row, 1, "Orders by Ship Mode", bg=GRAY)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        mode_df = df[s_col].value_counts().reset_index()
        mode_df.columns = ['Ship Mode', 'Count']
        mode_df['% Share'] = (mode_df['Count'] / len(df) * 100).round(1)
        if 'ship_cost' in cols:
            avg_cost = df.groupby(s_col)[cols['ship_cost']].mean().round(2)
            mode_df['Avg Cost'] = mode_df['Ship Mode'].map(avg_cost)
        if 'sales' in cols:
            avg_sale = df.groupby(s_col)[cols['sales']].mean().round(2)
            mode_df['Avg Sale'] = mode_df['Ship Mode'].map(avg_sale)

        headers = ['Ship Mode', 'Count', '% Share']
        if 'ship_cost' in cols: headers.append('Avg Cost')
        if 'sales' in cols: headers.append('Avg Sale')
        for ci, h in enumerate(headers, 1):
            hdr(ws, row, ci, h)
        row += 1

        mode_start = row
        for _, r in mode_df.iterrows():
            dat(ws, row, 1, r['Ship Mode'])
            dat(ws, row, 2, int(r['Count']), align="right")
            dat(ws, row, 3, f"{r['% Share']}%", align="right")
            ci = 4
            if 'ship_cost' in cols and 'Avg Cost' in mode_df.columns:
                cost = r.get('Avg Cost', 0) or 0
                flag_cell(ws, row, ci, f"${cost:.2f}", good=cost < 50)
                ci += 1
            if 'sales' in cols and 'Avg Sale' in mode_df.columns:
                dat(ws, row, ci, f"${r.get('Avg Sale', 0):.2f}", align="right")
            row += 1
        mode_end = row - 1

        d = Reference(ws, min_col=2, min_row=mode_start, max_row=mode_end)
        c = Reference(ws, min_col=1, min_row=mode_start, max_row=mode_end)
        pie_chart(ws, "Orders by Ship Mode", d, c, f"G{mode_start}")
        row += 2

    if 'ship_cost' in cols:
        sc = cols['ship_cost']
        hdr(ws, row, 1, "Shipping Cost Stats", bg=GRAY)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        stats = [
            ("Total Shipping Cost",       f"${df[sc].sum():,.2f}"),
            ("Average Shipping Cost",     f"${df[sc].mean():,.2f}"),
            ("Max Shipping Cost",         f"${df[sc].max():,.2f}"),
            ("Min Shipping Cost",         f"${df[sc].min():,.2f}"),
            ("Median Shipping Cost",      f"${df[sc].median():,.2f}"),
        ]
        if 'sales' in cols and df[cols['sales']].sum() > 0:
            pct = df[sc].sum() / df[cols['sales']].sum() * 100
            stats.append(("Shipping as % of Revenue", f"{pct:.1f}%"))

        for label, value in stats:
            dat(ws, row, 1, label, bold=True)
            dat(ws, row, 2, value, align="right")
            row += 1

        row += 1

        # Top 10 most expensive orders
        if 'product' in cols:
            hdr(ws, row, 1, "10 Most Expensive Shipping Orders", bg=GRAY)
            ws.merge_cells(f"A{row}:D{row}")
            row += 1
            expensive = df.nlargest(10, sc)[[cols['product'], sc] + ([cols['profit']] if 'profit' in cols else [])].reset_index(drop=True)
            for ci, col in enumerate(expensive.columns, 1):
                hdr(ws, row, ci, col)
            row += 1
            for _, r in expensive.iterrows():
                for ci, val in enumerate(r, 1):
                    if isinstance(val, float): val = round(val, 2)
                    dat(ws, row, ci, val)
                row += 1


# ─── SHEET 7: TIME SERIES (if date column exists) ─────────────────────────────
def build_timeseries(ws, df, cols):
    ws.sheet_view.showGridLines = False
    hdr(ws, 1, 1, "TIME SERIES ANALYSIS", size=13, bg=NAVY, fg=AMBER)
    ws.merge_cells("A1:F1")

    if 'date' not in cols:
        dat(ws, 3, 1, "No date column detected in dataset")
        return

    d_col = cols['date']
    df2 = df.copy()
    try:
        df2[d_col] = pd.to_datetime(df2[d_col], errors='coerce')
        df2 = df2.dropna(subset=[d_col])
        df2['_year']  = df2[d_col].dt.year
        df2['_month'] = df2[d_col].dt.to_period('M').astype(str)
    except Exception:
        dat(ws, 3, 1, "Could not parse date column")
        return

    set_widths(ws, {"A": 18, "B": 18, "C": 18, "D": 18})
    row = 3

    if 'sales' in cols:
        hdr(ws, row, 1, "Monthly Sales Trend", bg=GRAY)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        monthly = df2.groupby('_month').agg(
            **{'Total Sales': (cols['sales'], 'sum'),
               **({'Total Profit': (cols['profit'], 'sum')} if 'profit' in cols else {})}
        ).reset_index().sort_values('_month')
        monthly['Total Sales'] = monthly['Total Sales'].round(2)

        headers = ['Month', 'Total Sales']
        if 'profit' in cols: headers.append('Total Profit')
        for ci, h in enumerate(headers, 1):
            hdr(ws, row, ci, h)
        row += 1

        m_start = row
        for _, r in monthly.iterrows():
            dat(ws, row, 1, r['_month'])
            dat(ws, row, 2, r['Total Sales'], align="right")
            if 'profit' in cols:
                flag_cell(ws, row, 3, round(r['Total Profit'], 2), good=r['Total Profit'] > 0)
            row += 1
        m_end = row - 1

        d_ref = Reference(ws, min_col=2, min_row=m_start - 1, max_row=m_end)
        c_ref = Reference(ws, min_col=1, min_row=m_start, max_row=m_end)
        line_chart(ws, "Monthly Sales Trend", d_ref, c_ref, f"F{m_start - 1}", width=26, height=14)

    row += 2

    if 'sales' in cols:
        hdr(ws, row, 1, "Yearly Sales Summary", bg=GRAY)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        yearly = df2.groupby('_year').agg(
            **{'Total Sales': (cols['sales'], 'sum'),
               **({'Total Profit': (cols['profit'], 'sum')} if 'profit' in cols else {}),
               **({'Orders': (cols['order_id'], 'count')} if 'order_id' in cols else {})}
        ).reset_index()
        yearly['Total Sales'] = yearly['Total Sales'].round(2)

        for ci, h in enumerate(yearly.columns, 1):
            hdr(ws, row, ci, str(h))
        row += 1
        y_start = row
        for _, r in yearly.iterrows():
            for ci, val in enumerate(r, 1):
                if isinstance(val, float): val = round(val, 2)
                dat(ws, row, ci, val)
            row += 1
        y_end = row - 1

        d_ref = Reference(ws, min_col=2, min_row=y_start - 1, max_row=y_end)
        c_ref = Reference(ws, min_col=1, min_row=y_start, max_row=y_end)
        bar_chart(ws, "Yearly Sales", d_ref, c_ref, f"F{y_start}", width=18, height=12)


# ─── SHEET 8: FINAL VERDICT ───────────────────────────────────────────────────
def build_verdict(ws, df, cols, company_name):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 85

    hdr(ws, 1, 1, f"FINAL VERDICT — {company_name.upper()}", size=14, bg=NAVY, fg=AMBER)
    ws.row_dimensions[1].height = 30

    positives = []
    negatives = []

    if 'profit' in cols:
        total_profit = df[cols['profit']].sum()
        loss_rows = (df[cols['profit']] < 0).sum()
        loss_pct  = loss_rows / len(df) * 100

        if total_profit > 0:
            positives.append(f"Company is profitable overall — total profit: ${total_profit:,.2f}")
        else:
            negatives.append(f"Company is LOSS-MAKING — total loss: ${abs(total_profit):,.2f}")

        if loss_pct > 25:
            negatives.append(f"{loss_pct:.1f}% of transactions are loss-making ({loss_rows:,} rows out of {len(df):,})")
        elif loss_pct > 10:
            negatives.append(f"Concerning: {loss_pct:.1f}% of transactions run at a loss")
        else:
            positives.append(f"Low loss frequency — only {loss_pct:.1f}% of transactions are unprofitable")

        if 'sales' in cols and df[cols['sales']].sum() > 0:
            margin = total_profit / df[cols['sales']].sum() * 100
            if margin < 5:
                negatives.append(f"Critical: profit margin is only {margin:.1f}% — industry average is 10%+")
            elif margin < 10:
                negatives.append(f"Below-average margin of {margin:.1f}% — needs improvement")
            else:
                positives.append(f"Solid profit margin of {margin:.1f}%")

        if 'product' in cols:
            worst = df.groupby(cols['product'])[cols['profit']].sum().nsmallest(1)
            worst_name, worst_val = worst.index[0], worst.iloc[0]
            if worst_val < 0:
                negatives.append(f"Worst product '{worst_name}' destroys ${abs(worst_val):,.2f} in value")

    if 'ship_cost' in cols and 'sales' in cols:
        ship_pct = df[cols['ship_cost']].sum() / df[cols['sales']].sum() * 100
        if ship_pct > 15:
            negatives.append(f"Shipping costs consuming {ship_pct:.1f}% of revenue — critically high")
        elif ship_pct > 8:
            negatives.append(f"Shipping at {ship_pct:.1f}% of revenue — optimization needed")
        else:
            positives.append(f"Shipping costs controlled at {ship_pct:.1f}% of revenue")

    if 'region' in cols and 'sales' in cols:
        region_sales = df.groupby(cols['region'])[cols['sales']].sum()
        top_region = region_sales.idxmax()
        top_pct = region_sales.max() / region_sales.sum() * 100
        if top_pct > 60:
            negatives.append(f"Over-reliance on {top_region} ({top_pct:.0f}% of total sales) — concentration risk")
        else:
            positives.append(f"Good regional spread — top region ({top_region}) is {top_pct:.0f}% of sales")

    if 'discount' in cols:
        avg_disc = df[cols['discount']].mean()
        if avg_disc > 0.2:
            negatives.append(f"Average discount of {avg_disc*100:.1f}% is eroding margins")
        elif avg_disc > 0:
            positives.append(f"Controlled discounting — avg {avg_disc*100:.1f}%")

    # Score
    score = 50
    score += min(25, len(positives) * 6)
    score -= min(35, len(negatives) * 9)
    score = max(0, min(100, score))

    row = 3
    score_color = GREEN if score >= 65 else (AMBER if score >= 40 else RED)
    hdr(ws, row, 1, f"OVERALL SCORE: {score}/100", bg=score_color, fg=WHITE, size=14)
    ws.row_dimensions[row].height = 30
    row += 2

    hdr(ws, row, 1, "✅  SECTION 1 — POSITIVES", bg=GREEN, fg=WHITE, size=12)
    ws.row_dimensions[row].height = 24
    row += 1
    for p in (positives or ["No significant positives found"]):
        dat(ws, row, 1, f"✅   {p}")
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1
    hdr(ws, row, 1, "❌  SECTION 2 — NEGATIVES (WORST FIRST)", bg=RED, fg=WHITE, size=12)
    ws.row_dimensions[row].height = 24
    row += 1
    for n in (sorted(negatives, key=len, reverse=True) or ["No critical negatives found"]):
        dat(ws, row, 1, f"❌   {n}")
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1
    if score >= 70:
        verdict_text = "STRONG PERFORMER — Fundamentals are solid. Focus on scaling profitability."
    elif score >= 50:
        verdict_text = "AVERAGE — Operational improvements needed before aggressive expansion."
    elif score >= 30:
        verdict_text = "WEAK — Significant financial risks present. Immediate corrective action required."
    else:
        verdict_text = "CRITICAL — Company is in financial distress. Major restructuring is non-negotiable."

    hdr(ws, row, 1, f"⚖️   VERDICT: {verdict_text}", bg=NAVY, fg=AMBER, size=12)
    ws.row_dimensions[row].height = 28


# ─── SHEET 9: ACTION PLAN ─────────────────────────────────────────────────────
def build_action_plan(ws, df, cols):
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 5, "B": 58, "C": 14, "D": 18})

    hdr(ws, 1, 1, "ACTION PLAN", size=13, bg=NAVY, fg=AMBER)
    ws.merge_cells("A1:D1")

    row = 3
    hdr(ws, row, 1, "#")
    hdr(ws, row, 2, "ACTION ITEM")
    hdr(ws, row, 3, "PRIORITY")
    hdr(ws, row, 4, "TIMELINE")
    row += 1

    actions = []

    # Auto-generate from data
    if 'profit' in cols:
        loss_rows = (df[cols['profit']] < 0).sum()
        if loss_rows > 0:
            actions.append((f"Eliminate or reprice {loss_rows:,} loss-making transactions — immediate review required", "High", "2 weeks"))
        if 'product' in cols:
            worst3 = df.groupby(cols['product'])[cols['profit']].sum().nsmallest(3)
            for name, val in worst3.items():
                if val < 0:
                    actions.append((f"Discontinue or renegotiate pricing for: '{name}' (loss: ${abs(val):,.2f})", "High", "30 days"))

    if 'ship_cost' in cols and 'sales' in cols:
        pct = df[cols['ship_cost']].sum() / df[cols['sales']].sum() * 100
        if pct > 10:
            actions.append((f"Negotiate bulk shipping contracts — current rate {pct:.1f}% of revenue is unsustainable", "High", "60 days"))
            actions.append(("Shift non-urgent orders to Standard Class shipping to reduce freight costs", "Medium", "30 days"))

    if 'region' in cols and 'sales' in cols:
        bottom_region = df.groupby(cols['region'])[cols['sales']].sum().idxmin()
        actions.append((f"Investigate underperformance in {bottom_region} — develop region-specific strategy", "Medium", "90 days"))

    if 'discount' in cols and df[cols['discount']].mean() > 0.15:
        actions.append(("Cap maximum discount at 15% — current average discounting is eroding margins", "High", "Immediate"))

    actions += [
        ("Run monthly profitability review per product and region segment", "Medium", "Monthly"),
        ("Build real-time KPI dashboard tracking margin, shipping %, regional mix", "Low", "60 days"),
        ("Set minimum profit margin threshold of 15% for all new product launches", "Medium", "Next quarter"),
        ("Conduct quarterly review of bottom 10% performing products for discontinuation", "Low", "Quarterly"),
    ]

    priority_colors = {"High": RED, "Medium": AMBER, "Low": GREEN}
    for i, (action_text, priority, timeline) in enumerate(actions, start=1):
        dat(ws, row, 1, i, align="center")
        dat(ws, row, 2, action_text)
        c = ws.cell(row=row, column=3, value=priority)
        c.fill = PatternFill("solid", fgColor=priority_colors.get(priority, AMBER))
        c.font = Font(color=WHITE, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        dat(ws, row, 4, timeline)
        ws.row_dimensions[row].height = 22
        row += 1


# ─── MAIN GENERATOR ───────────────────────────────────────────────────────────
def generate_excel_report(
    company_name: str,
    df: pd.DataFrame = None,
    doc_type: str = "SALES_DATA",
    # Legacy params kept for backward compat — ignored if df is provided
    ratios: dict = None,
    positives: list = None,
    negatives: list = None,
    insights: list = None,
    actions: list = None,
    score: int = None,
    verdict: str = None,
    output_dir: str = "reports",
) -> str:
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{output_dir}/{company_name}_Reia_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if df is not None and not df.empty:
        # Disable charts for large datasets to avoid OOM
        global _CHARTS_ENABLED
        _CHARTS_ENABLED = len(df) <= 50000
        cols = detect_columns(df)

        build_overview(wb.create_sheet("Overview"), df, company_name, doc_type, cols)
        build_raw_data(wb.create_sheet("Raw Data"), df)

        if any(k in cols for k in ['sales', 'region', 'category', 'segment']):
            build_sales(wb.create_sheet("Sales Analysis"), df, cols)

        if 'product' in cols and 'sales' in cols:
            build_products(wb.create_sheet("Product Analysis"), df, cols)

        if 'profit' in cols:
            build_profit(wb.create_sheet("Profit Analysis"), df, cols)

        if 'ship_mode' in cols or 'ship_cost' in cols:
            build_shipping(wb.create_sheet("Shipping Analysis"), df, cols)

        if 'date' in cols:
            build_timeseries(wb.create_sheet("Time Series"), df, cols)

        build_verdict(wb.create_sheet("Final Verdict"), df, cols, company_name)
        build_action_plan(wb.create_sheet("Action Plan"), df, cols)

    else:
        ws = wb.create_sheet("Overview")
        ws.sheet_view.showGridLines = False
        hdr(ws, 1, 1, f"REIA Report — {company_name}", size=13)
        dat(ws, 3, 1, "No dataset provided for analysis. Upload a CSV/Excel file to generate a full report.")

    wb.save(filename)
    return filename
